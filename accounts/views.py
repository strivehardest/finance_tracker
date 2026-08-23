from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from datetime import datetime, timedelta
from decimal import Decimal
from .models import User, Account, Category, Transaction
from .forms import SignUpForm, TransactionForm, AccountForm, CategoryForm, ProfileForm
import requests
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.units import inch
from io import BytesIO

# Currency conversion helper
CURRENCY_SYMBOLS = {
    'GHS': '₵',
    'USD': '$',
    'EUR': '€',
    'GBP': '£',
    'NGN': '₦',
}

EXCHANGE_RATES_CACHE = {}
CACHE_TIME = None

def get_exchange_rates():
    """Get live exchange rates from API"""
    global EXCHANGE_RATES_CACHE, CACHE_TIME
    
    now = datetime.now()
    # Cache rates for 1 hour
    if CACHE_TIME and (now - CACHE_TIME).seconds < 3600 and EXCHANGE_RATES_CACHE:
        return EXCHANGE_RATES_CACHE
    
    try:
        # Using exchangerate-api.com free tier (GHS as base)
        response = requests.get('https://api.exchangerate-api.com/v4/latest/GHS', timeout=5)
        if response.status_code == 200:
            data = response.json()
            rates = data.get('rates', {})
            # Filter only needed currencies
            EXCHANGE_RATES_CACHE = {
                'GHS': 1.0,
                'USD': rates.get('USD', 0),
                'EUR': rates.get('EUR', 0),
                'GBP': rates.get('GBP', 0),
                'NGN': rates.get('NGN', 0),
            }
            CACHE_TIME = now
            return EXCHANGE_RATES_CACHE
    except:
        pass
    
    # Fallback rates if API fails (updated as of Nov 2025)
    return {
        'GHS': 1.0,
        'USD': 0.073,
        'EUR': 0.068,
        'GBP': 0.058,
        'NGN': 12.45,
    }

def convert_currency(amount, from_currency='GHS', to_currency='GHS'):
    """Convert amount from one currency to another"""
    if from_currency == to_currency:
        return amount
    
    rates = get_exchange_rates()
    
    # Convert to GHS first, then to target currency
    if from_currency != 'GHS':
        amount_ghs = amount / rates.get(from_currency, 1)
    else:
        amount_ghs = amount
    
    converted = amount_ghs * rates.get(to_currency, 1)
    return round(float(converted), 2)

def send_welcome_email(user):
    """Send welcome/verification email to new user"""
    try:
        subject = 'Welcome to Finance Tracker - Account Created Successfully'
        
        # Prepare email context
        context = {
            'first_name': user.first_name or user.username,
            'email': user.email,
            'username': user.username,
            'site_url': settings.SITE_URL if hasattr(settings, 'SITE_URL') else 'http://127.0.0.1:8000',
        }
        
        # Render HTML email template
        html_message = render_to_string('accounts/emails/welcome_email.html', context)
        plain_message = render_to_string('accounts/emails/welcome_email.txt', context)
        
        # Send email
        send_mail(
            subject=subject,
            message=plain_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            html_message=html_message,
            fail_silently=False,
        )
        return True
    except Exception as e:
        print(f"Error sending welcome email: {str(e)}")
        return False

def home(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'accounts/home.html')

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
        
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, f'Welcome back, {user.first_name or user.username}!')
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'accounts/login.html')

def signup_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
        
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            create_default_categories(user)
            Account.objects.create(
                name="Main Account",
                type="checking",
                balance=Decimal('0.00'),
                user=user
            )
            
            # Send welcome email
            email_sent = send_welcome_email(user)
            
            username = form.cleaned_data.get('username')
            if email_sent:
                messages.success(request, f'Account created for {username}! A welcome email has been sent to {user.email}')
            else:
                messages.success(request, f'Account created for {username}! You can now log in.')
                messages.warning(request, 'Note: Welcome email could not be sent. Please check your email settings.')
            
            return redirect('login')
    else:
        form = SignUpForm()
    
    return render(request, 'accounts/signup.html', {'form': form})

def logout_view(request):
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('home')

@login_required
def dashboard(request):
    user_currency = request.user.preferred_currency
    currency_symbol = CURRENCY_SYMBOLS.get(user_currency, '₵')
    now = datetime.now()

    accounts = Account.objects.filter(user=request.user, is_active=True)
    total_balance = accounts.aggregate(Sum('balance'))['balance__sum'] or Decimal('0.00')
    recent_transactions = Transaction.objects.filter(user=request.user)[:5]

    monthly_income = Transaction.objects.filter(
        user=request.user,
        date__month=now.month,
        date__year=now.year,
        category__type='income'
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')

    monthly_expenses = Transaction.objects.filter(
        user=request.user,
        date__month=now.month,
        date__year=now.year,
        category__type='expense'
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')

    total_income = Transaction.objects.filter(
        user=request.user,
        category__type='income'
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')

    total_expenses = Transaction.objects.filter(
        user=request.user,
        category__type='expense'
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')

    total_balance_converted = convert_currency(float(total_balance), 'GHS', user_currency)
    monthly_income_converted = convert_currency(float(monthly_income), 'GHS', user_currency)
    monthly_expenses_converted = convert_currency(float(monthly_expenses), 'GHS', user_currency)
    total_income_converted = convert_currency(float(total_income), 'GHS', user_currency)
    total_expenses_converted = convert_currency(float(total_expenses), 'GHS', user_currency)

    months_dict = {}
    year, month = now.year, now.month
    for _ in range(12):
        months_dict[datetime(year, month, 1).strftime('%b %Y')] = {'income': 0.0, 'expense': 0.0}
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    months_dict = dict(reversed(list(months_dict.items())))

    all_transactions = Transaction.objects.filter(user=request.user).values_list('date', 'category__type', 'amount')
    for trans_date, trans_type, trans_amount in all_transactions:
        month_key = trans_date.strftime('%b %Y')
        if month_key in months_dict:
            converted_amount = convert_currency(float(trans_amount), 'GHS', user_currency)
            if trans_type == 'income':
                months_dict[month_key]['income'] += converted_amount
            else:
                months_dict[month_key]['expense'] += converted_amount

    category_data = Transaction.objects.filter(
        user=request.user,
        date__month=now.month,
        date__year=now.year
    ).values('category__name').annotate(total=Sum('amount')).order_by('-total')

    category_data_list = []
    for item in category_data:
        converted_total = convert_currency(float(item['total']), 'GHS', user_currency) if item['total'] else 0
        category_data_list.append({
            'category__name': item['category__name'],
            'total': converted_total
        })

    for account in accounts:
        account.display_balance = convert_currency(float(account.balance), 'GHS', user_currency)
    for transaction in recent_transactions:
        transaction.display_amount = convert_currency(float(transaction.amount), 'GHS', user_currency)

    context = {
        'total_balance_converted': round(total_balance_converted, 2),
        'accounts': accounts,
        'recent_transactions': recent_transactions,
        'monthly_income_converted': round(monthly_income_converted, 2),
        'monthly_expenses_converted': round(monthly_expenses_converted, 2),
        'net_income_converted': round(monthly_income_converted - monthly_expenses_converted, 2),
        'total_income_converted': round(total_income_converted, 2),
        'total_expenses_converted': round(total_expenses_converted, 2),
        'net_balance_converted': round(total_income_converted - total_expenses_converted, 2),
        'currency_symbol': currency_symbol,
        'monthly_data': json.dumps(months_dict),
        'category_data': json.dumps(category_data_list),
    }

    return render(request, 'accounts/dashboard.html', context)

@login_required
def transactions_list(request):
    transactions = Transaction.objects.filter(user=request.user)
    
    category_filter = request.GET.get('category')
    if category_filter:
        transactions = transactions.filter(category_id=category_filter)
    
    account_filter = request.GET.get('account')
    if account_filter:
        transactions = transactions.filter(account_id=account_filter)
    
    search = request.GET.get('search')
    if search:
        transactions = transactions.filter(
            Q(description__icontains=search) | 
            Q(notes__icontains=search)
        )
    
    paginator = Paginator(transactions, 20)
    page = request.GET.get('page')
    transactions = paginator.get_page(page)
    
    context = {
        'transactions': transactions,
        'categories': Category.objects.filter(user=request.user),
        'accounts': Account.objects.filter(user=request.user, is_active=True),
        'current_category': category_filter,
        'current_account': account_filter,
        'search_term': search,
    }
    
    return render(request, 'accounts/transactions_list.html', context)

@login_required
def add_transaction(request):
    if request.method == 'POST':
        form = TransactionForm(request.POST, user=request.user)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.user = request.user
            transaction.save()
            messages.success(request, 'Transaction added successfully!')
            return redirect('transactions_list')
    else:
        form = TransactionForm(user=request.user)
    
    return render(request, 'accounts/add_transaction.html', {'form': form})

@login_required
def accounts_list(request):
    accounts = Account.objects.filter(user=request.user)
    return render(request, 'accounts/accounts_list.html', {'accounts': accounts})

@login_required
def add_account(request):
    if request.method == 'POST':
        form = AccountForm(request.POST)
        if form.is_valid():
            account = form.save(commit=False)
            account.user = request.user
            account.save()
            messages.success(request, 'Account added successfully!')
            return redirect('accounts_list')
    else:
        form = AccountForm()
    
    return render(request, 'accounts/add_account.html', {'form': form})

@login_required
def edit_account(request, id):
    account = get_object_or_404(Account, id=id, user=request.user)
    if request.method == 'POST':
        form = AccountForm(request.POST, instance=account)
        if form.is_valid():
            form.save()
            messages.success(request, 'Account updated successfully!')
            return redirect('accounts_list')
    else:
        form = AccountForm(instance=account)
    
    return render(request, 'accounts/edit_account.html', {'form': form, 'account': account})

@login_required
def delete_account(request, id):
    account = get_object_or_404(Account, id=id, user=request.user)
    if request.method == 'POST':
        account.delete()
        messages.success(request, 'Account deleted successfully!')
        return redirect('accounts_list')
    
    return render(request, 'accounts/delete_account.html', {'account': account})

@login_required
def categories_list(request):
    categories = Category.objects.filter(user=request.user)
    return render(request, 'accounts/categories_list.html', {'categories': categories})

@login_required
def add_category(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save(commit=False)
            category.user = request.user
            category.save()
            messages.success(request, 'Category added successfully!')
            return redirect('categories_list')
    else:
        form = CategoryForm()
    
    return render(request, 'accounts/add_category.html', {'form': form})

@login_required
def edit_category(request, pk):
    category = get_object_or_404(Category, pk=pk, user=request.user)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category updated successfully!')
            return redirect('categories_list')
    else:
        form = CategoryForm(instance=category)
    
    return render(request, 'accounts/edit_category.html', {'form': form, 'category': category})

@login_required
def delete_category(request, pk):
    category = get_object_or_404(Category, pk=pk, user=request.user)
    if request.method == 'POST':
        category.delete()
        messages.success(request, 'Category deleted successfully!')
        return redirect('categories_list')
    
    return render(request, 'accounts/delete_category.html', {'category': category})

def create_default_categories(user):
    default_categories = [
        {'name': 'Salary', 'type': 'income', 'icon': '💰', 'color': '#28a745'},
        {'name': 'Freelance', 'type': 'income', 'icon': '💻', 'color': '#17a2b8'},
        {'name': 'Food', 'type': 'expense', 'icon': '🍽️', 'color': '#fd7e14'},
        {'name': 'Transport', 'type': 'expense', 'icon': '🚗', 'color': '#6610f2'},
        {'name': 'Entertainment', 'type': 'expense', 'icon': '🎬', 'color': '#e83e8c'},
        {'name': 'Utilities', 'type': 'expense', 'icon': '⚡', 'color': '#dc3545'},
        {'name': 'Healthcare', 'type': 'expense', 'icon': '🏥', 'color': '#20c997'},
        {'name': 'Shopping', 'type': 'expense', 'icon': '🛍️', 'color': '#ffc107'},
    ]
    
    for cat_data in default_categories:
        Category.objects.create(user=user, **cat_data)

@login_required
def set_currency(request):
    """Update user's preferred currency"""
    if request.method == 'POST':
        currency = request.POST.get('currency', 'GHS')
        if currency in dict(User.CURRENCY_CHOICES):
            request.user.preferred_currency = currency
            request.user.save()
            messages.success(request, f'Currency changed to {currency}')
            return redirect('dashboard')
    
    return redirect('dashboard')

@login_required
def get_exchange_rates_api(request):
    """API endpoint to get current exchange rates"""
    rates = get_exchange_rates()
    return JsonResponse({
        'rates': rates,
        'base': 'GHS',
        'timestamp': datetime.now().isoformat()
    })

@login_required
def profile_view(request):
    """Display user profile"""
    # Calculate age if date of birth exists
    age = None
    if request.user.date_of_birth:
        today = datetime.now().date()
        age = today.year - request.user.date_of_birth.year - (
            (today.month, today.day) < (request.user.date_of_birth.month, request.user.date_of_birth.day)
        )
    
    context = {
        'user': request.user,
        'has_profile_picture': bool(request.user.profile_picture),
        'age': age,
    }
    return render(request, 'accounts/profile.html', context)

@login_required
def edit_profile(request):
    """Edit user profile information"""
    if request.method == 'POST':
        form = ProfileForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            user = form.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('profile')
        else:
            # Display form errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ProfileForm(instance=request.user)
    
    context = {
        'form': form,
        'user': request.user,
        'has_profile_picture': bool(request.user.profile_picture),
    }
    return render(request, 'accounts/edit_profile.html', context)

@login_required
def export_transactions_excel(request):
    """Export transactions to Excel file"""
    # Get all transactions for the user
    transactions = Transaction.objects.filter(user=request.user).order_by('-date')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    # Define styles
    header_fill = PatternFill(start_color="f57c00", end_color="f57c00", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers
    headers = ['Date', 'Description', 'Category', 'Type', 'Amount', 'Account', 'Notes']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    for row_num, transaction in enumerate(transactions, 2):
        ws.cell(row=row_num, column=1).value = transaction.date.strftime('%Y-%m-%d')
        ws.cell(row=row_num, column=2).value = transaction.description
        ws.cell(row=row_num, column=3).value = transaction.category.name
        ws.cell(row=row_num, column=4).value = transaction.category.type.capitalize()
        ws.cell(row=row_num, column=5).value = float(transaction.amount)
        ws.cell(row=row_num, column=6).value = transaction.account.name
        ws.cell(row=row_num, column=7).value = transaction.notes
        
        # Apply formatting
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            if col == 5:  # Amount column
                cell.number_format = '#,##0.00'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20
    
    # Add summary
    summary_row = len(transactions) + 3
    ws.cell(row=summary_row, column=1).value = "Summary"
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=11)
    
    total_income = transactions.filter(category__type='income').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    total_expenses = transactions.filter(category__type='expense').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    
    ws.cell(row=summary_row + 1, column=1).value = "Total Income:"
    ws.cell(row=summary_row + 1, column=2).value = float(total_income)
    ws.cell(row=summary_row + 1, column=2).font = Font(color="008000", bold=True)
    ws.cell(row=summary_row + 1, column=2).number_format = '#,##0.00'
    
    ws.cell(row=summary_row + 2, column=1).value = "Total Expenses:"
    ws.cell(row=summary_row + 2, column=2).value = float(total_expenses)
    ws.cell(row=summary_row + 2, column=2).font = Font(color="FF0000", bold=True)
    ws.cell(row=summary_row + 2, column=2).number_format = '#,##0.00'
    
    ws.cell(row=summary_row + 3, column=1).value = "Net Balance:"
    ws.cell(row=summary_row + 3, column=2).value = float(total_income - total_expenses)
    ws.cell(row=summary_row + 3, column=2).font = Font(bold=True, size=11)
    ws.cell(row=summary_row + 3, column=2).number_format = '#,##0.00'
    
    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="transactions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response

@login_required
def export_transactions_pdf(request):
    """Export transactions to PDF file"""
    # Get all transactions for the user
    transactions = Transaction.objects.filter(user=request.user).order_by('-date')
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#f57c00'),
        spaceAfter=12,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.grey,
        spaceAfter=12
    )
    
    # Add title
    elements.append(Paragraph("Transaction Report", title_style))
    elements.append(Paragraph(f"Generated on {datetime.now().strftime('%B %d, %Y at %H:%M')}", subtitle_style))
    elements.append(Paragraph(f"User: {request.user.get_full_name() or request.user.username}", subtitle_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Prepare table data
    data = [['Date', 'Description', 'Category', 'Type', 'Amount', 'Account']]
    
    for transaction in transactions:
        data.append([
            transaction.date.strftime('%Y-%m-%d'),
            transaction.description[:25],  # Truncate for PDF
            transaction.category.name,
            transaction.category.type.capitalize(),
            f"₵{float(transaction.amount):.2f}",
            transaction.account.name[:20]  # Truncate for PDF
        ])
    
    # Create table
    table = Table(data, colWidths=[1*inch, 1.5*inch, 1.2*inch, 0.8*inch, 1*inch, 1.5*inch])
    
    # Style table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f57c00')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Add summary
    total_income = transactions.filter(category__type='income').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    total_expenses = transactions.filter(category__type='expense').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    
    summary_data = [
        ['Total Income', f"₵{float(total_income):.2f}"],
        ['Total Expenses', f"₵{float(total_expenses):.2f}"],
        ['Net Balance', f"₵{float(total_income - total_expenses):.2f}"],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    
    elements.append(Paragraph("Summary", styles['Heading3']))
    elements.append(summary_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="transactions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response


@login_required
def budget_view(request):
    """Budget management view"""
    # Get only expense categories with budget limits
    categories = Category.objects.filter(user=request.user, type='expense')
    
    # Calculate spending for each category this month
    current_month_start = datetime.now().replace(day=1)
    next_month = current_month_start + timedelta(days=32)
    next_month_start = next_month.replace(day=1)
    
    budget_data = []
    total_budget = Decimal('0.00')
    total_spent = Decimal('0.00')
    
    for category in categories:
        # Only include categories with budget limits
        if not category.budget_limit or category.budget_limit == 0:
            continue
            
        # Get spending for this month
        spent = Transaction.objects.filter(
            user=request.user,
            category=category,
            date__gte=current_month_start,
            date__lt=next_month_start
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        budget_limit = Decimal(str(category.budget_limit)) if category.budget_limit else Decimal('0.00')
        remaining = budget_limit - spent if budget_limit > 0 else Decimal('0.00')
        
        budget_percent = 0
        if budget_limit > 0:
            budget_percent = int((spent / budget_limit) * 100)
        
        total_budget += budget_limit
        total_spent += spent
        
        budget_data.append({
            'id': category.id,
            'name': category.name,
            'icon': category.icon,
            'budget_limit': f"{request.user.preferred_currency} {budget_limit:.2f}",
            'spent_this_month': f"{request.user.preferred_currency} {spent:.2f}",
            'remaining_budget': f"{request.user.preferred_currency} {remaining:.2f}",
            'budget_percent': budget_percent,
        })
    
    # Calculate overall budget usage
    total_budget_percent = 0
    if total_budget > 0:
        total_budget_percent = int((total_spent / total_budget) * 100)
    
    context = {
        'categories': budget_data,
        'total_budget': f"{request.user.preferred_currency} {total_budget:.2f}",
        'total_spent': f"{request.user.preferred_currency} {total_spent:.2f}",
        'budget_percent': total_budget_percent,
    }
    
    return render(request, 'accounts/budget.html', context)


@login_required
def edit_transaction(request, id):
    transaction = get_object_or_404(Transaction, id=id, user=request.user)
    if request.method == 'POST':
        form = TransactionForm(request.POST, instance=transaction, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Transaction updated successfully!')
            return redirect('transactions_list')
    else:
        form = TransactionForm(instance=transaction, user=request.user)
    
    return render(request, 'accounts/edit_transaction.html', {'form': form, 'transaction': transaction})


@login_required
def delete_transaction(request, id):
    transaction = get_object_or_404(Transaction, id=id, user=request.user)
    if request.method == 'POST':
        transaction.delete()
        messages.success(request, 'Transaction deleted successfully!')
        return redirect('transactions_list')
    
    return render(request, 'accounts/delete_transaction.html', {'transaction': transaction})
