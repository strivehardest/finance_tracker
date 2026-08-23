from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO

from django.conf import settings
from django.core.files.uploadedfile import InMemoryUploadedFile, UploadedFile
from django.core.mail import EmailMessage
from django.db.models import Sum
from django.utils.html import escape
from django.utils.safestring import mark_safe
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

CURRENCY_SYMBOLS = {
    'GHS': '₵',
    'USD': '$',
    'EUR': '€',
    'GBP': '£',
    'NGN': '₦',
}

EMOJI_TO_FA = {
    '💰': 'fa-money-bill-wave',
    '💵': 'fa-money-bill-wave',
    '💸': 'fa-money-bill-wave',
    '💻': 'fa-laptop',
    '🖥️': 'fa-laptop',
    '🏢': 'fa-building',
    '📈': 'fa-chart-line',
    '🎁': 'fa-gift',
    '🍔': 'fa-utensils',
    '🍕': 'fa-utensils',
    '🍽️': 'fa-utensils',
    '☕': 'fa-coffee',
    '🚗': 'fa-car',
    '🚌': 'fa-bus',
    '⛽': 'fa-gas-pump',
    '✈️': 'fa-plane',
    '🏠': 'fa-home',
    '💡': 'fa-bolt',
    '🔌': 'fa-bolt',
    '📶': 'fa-wifi',
    '📱': 'fa-mobile-alt',
    '🎬': 'fa-film',
    '🎮': 'fa-gamepad',
    '🎵': 'fa-music',
    '❤️': 'fa-heartbeat',
    '🏥': 'fa-heartbeat',
    '💊': 'fa-pills',
    '🛍️': 'fa-shopping-bag',
    '👕': 'fa-tshirt',
    '🎓': 'fa-graduation-cap',
    '🐾': 'fa-paw',
    '📦': 'fa-box',
    '💳': 'fa-credit-card',
}


def normalize_icon(icon):
    icon = (icon or '').strip()
    if icon.startswith('fa-'):
        return icon
    return EMOJI_TO_FA.get(icon, 'fa-tag')


def pagination_pages(page_obj, adjacent=1):
    if not page_obj or not page_obj.paginator.num_pages:
        return []
    current = page_obj.number
    last = page_obj.paginator.num_pages
    nums = {1, last}
    for number in range(current - adjacent, current + adjacent + 1):
        if 1 <= number <= last:
            nums.add(number)
    result = []
    previous = 0
    for number in sorted(nums):
        if previous and number > previous + 1:
            result.append(None)
        result.append(number)
        previous = number
    return result


def process_profile_image(picture):
    picture.seek(0)
    image = Image.open(picture)
    image = ImageOps.exif_transpose(image)
    if image.mode in ('RGBA', 'LA', 'P'):
        rgba = image.convert('RGBA')
        background = Image.new('RGB', rgba.size, (255, 255, 255))
        background.paste(rgba, mask=rgba.split()[-1])
        image = background
    else:
        image = image.convert('RGB')
    image = ImageOps.fit(image, (512, 512), Image.Resampling.LANCZOS)
    buffer = BytesIO()
    image.save(buffer, format='JPEG', quality=86, optimize=True)
    buffer.seek(0)
    original = getattr(picture, 'name', 'profile.jpg') or 'profile.jpg'
    stem = original.rsplit('/', 1)[-1].rsplit('.', 1)[0] or 'profile'
    return InMemoryUploadedFile(
        buffer,
        'ImageField',
        f'{stem}.jpg',
        'image/jpeg',
        buffer.getbuffer().nbytes,
        None,
    )


def is_uploaded_file(value):
    return isinstance(value, UploadedFile)


PERIOD_CHOICES = (
    ('this_month', 'This month'),
    ('last_month', 'Last month'),
    ('last_30', 'Last 30 days'),
    ('last_90', 'Last 90 days'),
    ('this_year', 'This year'),
    ('all', 'All time'),
    ('custom', 'Custom dates'),
)


def profile_photo_url(user):
    if not user or not getattr(user, 'is_authenticated', False):
        return ''
    picture = getattr(user, 'profile_picture', None)
    if not picture:
        return ''
    try:
        if picture.name:
            return picture.url
    except Exception:
        return ''
    return ''


def category_icon_html(icon, color='#f57c00'):
    icon = normalize_icon(icon)
    color = escape(color or '#f57c00')
    return mark_safe(
        f'<span class="cat-icon" style="--cat:{color}"><i class="fas {escape(icon)}"></i></span>'
    )


def parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except (TypeError, ValueError):
        return None


def resolve_period(period, start_str=None, end_str=None):
    today = datetime.now().date()
    period = period or 'all'
    if period == 'this_month':
        return today.replace(day=1), today, 'This month'
    if period == 'last_month':
        first = today.replace(day=1)
        last_end = first - timedelta(days=1)
        return last_end.replace(day=1), last_end, 'Last month'
    if period == 'last_30':
        return today - timedelta(days=29), today, 'Last 30 days'
    if period == 'last_90':
        return today - timedelta(days=89), today, 'Last 90 days'
    if period == 'this_year':
        return today.replace(month=1, day=1), today, 'This year'
    if period == 'custom':
        start = parse_date(start_str)
        end = parse_date(end_str)
        if start and end:
            if start > end:
                start, end = end, start
            return start, end, f'{start:%d %b %Y} – {end:%d %b %Y}'
    return None, None, 'All time'


def filter_transactions_period(queryset, period, start_str=None, end_str=None):
    start, end, label = resolve_period(period, start_str, end_str)
    if start:
        queryset = queryset.filter(date__gte=start)
    if end:
        queryset = queryset.filter(date__lte=end)
    return queryset, start, end, label


def _summaries(transactions):
    income = transactions.filter(category__type='income').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    expenses = transactions.filter(category__type='expense').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    return income, expenses, income - expenses


def _currency_code(user):
    return getattr(user, 'preferred_currency', None) or 'GHS'


def build_excel_report(user, transactions, period_label):
    income, expenses, net = _summaries(transactions)
    currency = _currency_code(user)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Transactions'

    orange = PatternFill('solid', fgColor='F57C00')
    ink = PatternFill('solid', fgColor='1C1917')
    zebra = PatternFill('solid', fgColor='FFF7ED')
    green = Font(color='15803D', bold=True)
    red = Font(color='B91C1C', bold=True)
    white = Font(bold=True, color='FFFFFF', size=12)
    title = Font(bold=True, color='FFFFFF', size=16)
    thin = Border(
        left=Side(style='thin', color='E7E5E4'),
        right=Side(style='thin', color='E7E5E4'),
        top=Side(style='thin', color='E7E5E4'),
        bottom=Side(style='thin', color='E7E5E4'),
    )

    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = 'Finance Tracker'
    cell.fill = ink
    cell.font = title
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:G2')
    ws['A2'].value = (
        f'Transaction report · {period_label} · {currency} · '
        f'{user.get_full_name() or user.username}'
    )
    ws['A2'].font = Font(color='78716C', italic=True)

    headers = ['Date', 'Description', 'Category', 'Type', f'Amount ({currency})', 'Account', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = orange
        cell.font = white
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin

    for row_num, transaction in enumerate(transactions, 5):
        values = [
            transaction.date.strftime('%Y-%m-%d'),
            transaction.description,
            transaction.category.name,
            transaction.category.type.title(),
            float(transaction.amount),
            transaction.account.name,
            transaction.notes or '',
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin
            if row_num % 2 == 0:
                cell.fill = zebra
            if col == 5:
                cell.number_format = '#,##0.00'
                cell.font = green if transaction.category.type == 'income' else red

    last_data = 4 + transactions.count()
    ws.auto_filter.ref = f'A4:G{max(last_data, 4)}'
    ws.freeze_panes = 'A5'
    widths = [14, 28, 16, 12, 14, 18, 28]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    summary_row = last_data + 2
    ws.cell(row=summary_row, column=1, value='Summary').font = Font(bold=True, size=12)
    rows = (
        (f'Total income ({currency})', float(income), green),
        (f'Total expenses ({currency})', float(expenses), red),
        (f'Net ({currency})', float(net), Font(bold=True)),
    )
    for offset, (label, value, font) in enumerate(rows, 1):
        ws.cell(row=summary_row + offset, column=1, value=label)
        amount = ws.cell(row=summary_row + offset, column=2, value=value)
        amount.number_format = '#,##0.00'
        amount.font = font

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_pdf_report(user, transactions, period_label):
    income, expenses, net = _summaries(transactions)
    currency = _currency_code(user)
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=0.7 * inch,
        bottomMargin=0.6 * inch,
        title='Finance Tracker report',
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'ReportTitle', parent=styles['Heading1'], fontSize=18,
        textColor=colors.HexColor('#1c1917'), spaceAfter=2, fontName='Helvetica-Bold',
    )
    meta_style = ParagraphStyle(
        'ReportMeta', parent=styles['Normal'], fontSize=9,
        textColor=colors.HexColor('#78716c'), spaceAfter=10,
    )
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8, leading=11)

    story = [
        Paragraph('Finance Tracker', title_style),
        Paragraph(
            f'{period_label} · {currency} · {user.get_full_name() or user.username} · generated {datetime.now():%d %b %Y %H:%M}',
            meta_style,
        ),
    ]

    header = ['Date', 'Description', 'Category', 'Type', f'Amount ({currency})', 'Account']
    data = [header]
    for transaction in transactions:
        amount = f"{'+' if transaction.category.type == 'income' else '-'}{float(transaction.amount):,.2f}"
        data.append([
            transaction.date.strftime('%d %b %Y'),
            Paragraph(escape(transaction.description[:80]), cell_style),
            transaction.category.name,
            transaction.category.type.title(),
            amount,
            transaction.account.name,
        ])

    table = Table(data, colWidths=[1.15*inch, 3.1*inch, 1.4*inch, 0.9*inch, 1.1*inch, 1.5*inch], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f57c00')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#e7e5e4')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fff7ed')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(table)
    story.append(Spacer(1, 0.25 * inch))

    summary = Table(
        [
            [f'Total income ({currency})', f'{float(income):,.2f}'],
            [f'Total expenses ({currency})', f'{float(expenses):,.2f}'],
            [f'Net ({currency})', f'{float(net):,.2f}'],
        ],
        colWidths=[2.2 * inch, 1.4 * inch],
    )
    summary.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('TEXTCOLOR', (1, 0), (1, 0), colors.HexColor('#15803d')),
        ('TEXTCOLOR', (1, 1), (1, 1), colors.HexColor('#b91c1c')),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#fff7ed')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#f57c00')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(summary)

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setFillColor(colors.HexColor('#f57c00'))
        canvas.rect(0, landscape(A4)[1] - 18, landscape(A4)[0], 18, fill=1, stroke=0)
        canvas.setFillColor(colors.white)
        canvas.setFont('Helvetica', 8)
        canvas.drawString(36, landscape(A4)[1] - 13, 'Finance Tracker')
        canvas.drawRightString(landscape(A4)[0] - 36, 20, f'Page {doc_.page}')
        canvas.restoreState()

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


def email_report(user, filename, content, mimetype, period_label, to_email):
    message = EmailMessage(
        subject=f'Finance Tracker report — {period_label}',
        body=(
            f'Hi {user.first_name or user.username},\n\n'
            f'Attached is your transaction report ({period_label}).\n\n'
            'Finance Tracker'
        ),
        from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', None),
        to=[to_email],
    )
    message.attach(filename, content, mimetype)
    message.send()
